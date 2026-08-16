import logging
import threading
from dataclasses import dataclass
from io import BytesIO
from re import search
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.learning_item import LearningItem
from app.schemas.learning import ImportSkippedItem
from app.services.llm_translation import LlmTranslationSettings, enrich_word_with_phonetics, generate_phonetic_decomposition, needs_translation, translate_english_to_chinese
from app.services.speech_asset_cache import precache_learning_speech_assets
from app.services.word_translation_cache import ensure_word_translations, extract_unique_words, sanitize_word_translation

SUPPORTED_IMPORT_EXTENSIONS = {".txt", ".xlsx"}

logger = logging.getLogger("learning_import")

# 同一用户的并发导入会同时通过各自的 in-memory 去重，插入重复的
# learning_items（表上没有唯一约束——存量数据可能已有重复，不能补约束）。
# 用 per-user 锁串行化整个导入。该方案成立的前提是生产只跑单个 uvicorn
# worker，进程内锁足够；若将来改为多 worker，必须换成 DB 级锁/约束。
_import_locks: dict[UUID, threading.Lock] = {}
_import_locks_guard = threading.Lock()


def _get_import_lock(user_id: UUID) -> threading.Lock:
    with _import_locks_guard:
        lock = _import_locks.get(user_id)
        if lock is None:
            lock = threading.Lock()
            _import_locks[user_id] = lock
        return lock


@dataclass(frozen=True)
class ParsedLearningItem:
    item_type: str
    english_text: str
    chinese_text: str
    phonetic: str | None = None
    difficulty_level: int = 1
    source: str | None = None


@dataclass(frozen=True)
class ImportParseResult:
    items: list[ParsedLearningItem]
    skipped_items: list[ImportSkippedItem]
    total_rows: int


def classify_learning_item(english_text: str) -> str:
    normalized = english_text.strip()
    word_count = len([part for part in normalized.replace("-", " ").split() if part])
    has_sentence_mark = any(mark in normalized for mark in (".", "?", "!"))
    has_finite_verb_hint = bool(search(r"\b(am|is|are|was|were|do|does|did|have|has|had|can|will|like|likes|go|goes|went|see|sees|want|wants)\b", normalized.lower()))

    if word_count <= 1 and not has_sentence_mark:
        return "word"
    if word_count >= 4 or has_sentence_mark or has_finite_verb_hint:
        return "sentence"
    return "phrase"


def contains_chinese(value: str) -> bool:
    return any("一" <= character <= "鿿" for character in value)


def split_learning_line(stripped_line: str) -> tuple[str, str]:
    for separator in ("\t", "|", "：", ":"):
        if separator in stripped_line:
            english_text, chinese_text = stripped_line.split(separator, 1)
            return english_text.strip(), chinese_text.strip()

    if "," in stripped_line:
        english_text, possible_chinese_text = stripped_line.rsplit(",", 1)
        if contains_chinese(possible_chinese_text):
            return english_text.strip(), possible_chinese_text.strip()

    return stripped_line, ""


def parse_learning_line(line: str, source: str) -> ParsedLearningItem | None:
    stripped_line = line.strip()
    if not stripped_line:
        return None

    english_text, chinese_text = split_learning_line(stripped_line)

    return ParsedLearningItem(
        item_type=classify_learning_item(english_text),
        english_text=english_text,
        chinese_text=chinese_text,
        source=source,
    )


def parse_txt_import(content: bytes, filename: str) -> ImportParseResult:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            # Common for files exported from Chinese Windows tools
            text = content.decode("gb18030")
        except UnicodeDecodeError as exc:
            raise ValueError("文件编码无法识别，请使用 UTF-8 或 GBK 编码的文本文件") from exc
    items: list[ParsedLearningItem] = []
    skipped_items: list[ImportSkippedItem] = []
    rows = text.splitlines()

    for row in rows:
        parsed_item = parse_learning_line(row, filename)
        if parsed_item is None:
            continue
        if not parsed_item.english_text.strip():
            skipped_items.append(ImportSkippedItem(english_text="", reason="English text is empty"))
            continue
        items.append(parsed_item)

    return ImportParseResult(items=items, skipped_items=skipped_items, total_rows=len(rows))


def parse_xlsx_import(content: bytes, filename: str) -> ImportParseResult:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    items: list[ParsedLearningItem] = []
    skipped_items: list[ImportSkippedItem] = []
    total_rows = 0

    for row in worksheet.iter_rows(values_only=True):
        total_rows += 1
        values = [str(value).strip() if value is not None else "" for value in row]
        if not any(values):
            continue

        first_cell = values[0].lower()
        if total_rows == 1 and first_cell in {"english", "english_text", "英文", "英语"}:
            continue

        english_text = values[0]
        chinese_text = values[1] if len(values) > 1 and values[1] else ""
        item_type = values[2].lower() if len(values) > 2 and values[2].lower() in {"word", "phrase", "sentence"} else classify_learning_item(english_text)
        phonetic = values[3] if len(values) > 3 and values[3] else None
        difficulty_level = parse_difficulty_level(values[4] if len(values) > 4 else "")

        if not english_text:
            skipped_items.append(ImportSkippedItem(english_text="", reason="English text is empty"))
            continue

        items.append(
            ParsedLearningItem(
                item_type=item_type,
                english_text=english_text,
                chinese_text=chinese_text,
                phonetic=phonetic,
                difficulty_level=difficulty_level,
                source=filename,
            )
        )

    return ImportParseResult(items=items, skipped_items=skipped_items, total_rows=total_rows)


def parse_difficulty_level(value: str) -> int:
    if not value:
        return 1
    try:
        parsed_value = int(value)
    except ValueError:
        return 1
    return min(max(parsed_value, 1), 5)


def normalize_english_text(value: str) -> str:
    return " ".join(value.strip().lower().split())


def import_learning_items(
    db: Session,
    user_id: UUID,
    course_id: UUID,
    parsed_items: list[ParsedLearningItem],
    translation_settings: LlmTranslationSettings | None = None,
    stored_settings: dict[str, object] | None = None,
) -> tuple[list[LearningItem], list[ImportSkippedItem]]:
    # 持锁贯穿整个导入（含 LLM 翻译/precache 等慢步骤），理由见上方锁注释。
    with _get_import_lock(user_id):
        return _import_learning_items_locked(
            db, user_id, course_id, parsed_items, translation_settings, stored_settings,
        )


def _import_learning_items_locked(
    db: Session,
    user_id: UUID,
    course_id: UUID,
    parsed_items: list[ParsedLearningItem],
    translation_settings: LlmTranslationSettings | None = None,
    stored_settings: dict[str, object] | None = None,
) -> tuple[list[LearningItem], list[ImportSkippedItem]]:
    imported_items: list[LearningItem] = []
    skipped_items: list[ImportSkippedItem] = []
    seen_keys: set[tuple[str, str]] = set()
    translation_failure_reason: str | None = None

    existing_rows = db.execute(
        select(func.lower(LearningItem.english_text), LearningItem.item_type).where(
            LearningItem.user_id == user_id,
            LearningItem.course_id == course_id,
        )
    ).all()
    existing_keys = {(normalize_english_text(row[0]), row[1]) for row in existing_rows}

    for parsed_item in parsed_items:
        normalized_english = normalize_english_text(parsed_item.english_text)
        item_key = (normalized_english, parsed_item.item_type)
        if item_key in seen_keys:
            skipped_items.append(ImportSkippedItem(english_text=parsed_item.english_text, reason="Duplicate in uploaded file"))
            continue
        if item_key in existing_keys:
            skipped_items.append(ImportSkippedItem(english_text=parsed_item.english_text, reason="Already exists"))
            continue

        chinese_text = parsed_item.chinese_text.strip()
        if needs_translation(chinese_text):
            if translation_settings is None:
                skipped_items.append(ImportSkippedItem(english_text=parsed_item.english_text, reason="缺少中文释义，且未配置可用的 LLM 翻译服务"))
                continue
            if translation_failure_reason is not None:
                skipped_items.append(ImportSkippedItem(english_text=parsed_item.english_text, reason=translation_failure_reason))
                continue
            try:
                if parsed_item.item_type == "word":
                    # New words get 1-3 common Chinese meanings from the LLM, not just one.
                    chinese_text = sanitize_word_translation(
                        translate_english_to_chinese(parsed_item.english_text, translation_settings, multiple_meanings=True),
                        source_word=parsed_item.english_text,
                    )
                    if not chinese_text:
                        raise ValueError("LLM translation failed: empty after sanitize")
                else:
                    chinese_text = translate_english_to_chinese(parsed_item.english_text, translation_settings)
            except ValueError as exc:
                translation_failure_reason = str(exc)
                skipped_items.append(ImportSkippedItem(english_text=parsed_item.english_text, reason=translation_failure_reason))
                continue

        learning_item = LearningItem(
            user_id=user_id,
            course_id=course_id,
            item_type=parsed_item.item_type,
            english_text=parsed_item.english_text.strip(),
            chinese_text=chinese_text,
            phonetic=parsed_item.phonetic,
            difficulty_level=parsed_item.difficulty_level,
            source=parsed_item.source,
        )
        db.add(learning_item)
        imported_items.append(learning_item)
        seen_keys.add(item_key)

    db.commit()
    for item in imported_items:
        db.refresh(item)

    # 重试补偿（2026-08-16）：items commit 之后还有翻译、音标 enrichment、
    # 词翻译缓存、语音 precache 多步；此前任一步以非 ValueError 失败都会
    # 留下"条目已建但 enrichment 未做"的半成品，且重试时这些条目全部按
    # "Already exists" 被跳过、永远不会补齐。这里把本次 parsed 命中的
    # 已存在条目捞回来：缺 enrichment 产物的交给下面的步骤补做，
    # precache 对全部命中条目幂等重跑。
    parsed_keys = {
        (normalize_english_text(parsed_item.english_text), parsed_item.item_type)
        for parsed_item in parsed_items
    }
    imported_ids = {item.id for item in imported_items}
    retry_items: list[LearningItem] = []
    if parsed_keys:
        for item in db.scalars(
            select(LearningItem).where(
                LearningItem.user_id == user_id,
                LearningItem.course_id == course_id,
            )
        ).all():
            if item.id in imported_ids:
                continue
            if (normalize_english_text(item.english_text), item.item_type) in parsed_keys:
                retry_items.append(item)

    # items 已提交，后续每一步的失败只允许降级为部分成功响应，不允许再把
    # 整个请求 500 掉——统一 except Exception + logger.exception。

    # 补齐中文释义：重试命中但缺中文的已有条目在这里翻译，不重新建条目。
    if translation_settings is not None:
        translated_any = False
        for item in retry_items:
            if not needs_translation(item.chinese_text or ""):
                continue
            try:
                if item.item_type == "word":
                    chinese_text = sanitize_word_translation(
                        translate_english_to_chinese(item.english_text, translation_settings, multiple_meanings=True),
                        source_word=item.english_text,
                    )
                    if not chinese_text:
                        raise ValueError("LLM translation failed: empty after sanitize")
                else:
                    chinese_text = translate_english_to_chinese(item.english_text, translation_settings)
                item.chinese_text = chinese_text
                translated_any = True
            except Exception:
                logger.exception("Import retry translation failed: english_text=%s", item.english_text)
        if translated_any:
            try:
                db.commit()
            except Exception:
                db.rollback()
                logger.exception("Import retry translation commit failed: user_id=%s course_id=%s", user_id, course_id)

    # Enrich word-type items with phonetic data (syllables + grapheme-phoneme map + IPA)
    if translation_settings is not None:
        word_items = [item for item in imported_items if item.item_type == "word"]
        # 重试命中的 word 条目缺音节数据时一并补做（syllables 与 enrichment
        # 同事务写入，缺失即代表上次 enrichment 未完成）
        word_items.extend(item for item in retry_items if item.item_type == "word" and not item.syllables)
        for item in word_items:
            try:
                syllables, gp_map = enrich_word_with_phonetics(item.english_text, translation_settings)
                item.syllables = syllables
                item.grapheme_phoneme_map = gp_map
                # Also generate IPA if not already provided
                if not item.phonetic:
                    try:
                        decomposition = generate_phonetic_decomposition(item.english_text, translation_settings)
                        if decomposition.get("ipa"):
                            item.phonetic = str(decomposition["ipa"])
                    except Exception:
                        logger.exception("IPA decomposition failed during import: english_text=%s", item.english_text)
            except Exception:
                logger.exception("Phonetic enrichment failed during import: english_text=%s", item.english_text)
        if word_items:
            try:
                db.commit()
                for item in word_items:
                    db.refresh(item)
            except Exception:
                db.rollback()
                logger.exception("Phonetic enrichment commit failed during import: user_id=%s course_id=%s", user_id, course_id)

        term_source_items = imported_items + retry_items
        course_terms = extract_unique_words([item.english_text for item in term_source_items])
        course_terms.extend(
            normalize_english_text(item.english_text)
            for item in term_source_items
            if item.item_type in {"word", "phrase"} and item.english_text.strip()
        )
        if course_terms:
            try:
                ensure_word_translations(db, user_id, course_terms, translation_settings, course_id)
                db.commit()
            except Exception:
                db.rollback()
                logger.exception("Word translation cache failed during import: user_id=%s course_id=%s", user_id, course_id)

    # 语音 precache 幂等（命中缓存即跳过），对本次导入命中的全部条目重跑，
    # 保证上次在 precache 之前/之中失败时重试能补完。
    precache_items = imported_items + retry_items
    if precache_items:
        try:
            precache_learning_speech_assets(
                db,
                user_id=user_id,
                course_id=course_id,
                learning_items=precache_items,
                stored_settings=stored_settings,
            )
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("Speech precache failed during import: user_id=%s course_id=%s", user_id, course_id)

    # Assign sort order after import
    try:
        resequence_course_items(db, user_id, course_id)
    except Exception:
        db.rollback()
        logger.exception("Resequence failed during import: user_id=%s course_id=%s", user_id, course_id)

    return imported_items, skipped_items


def _edit_distance(a: str, b: str) -> int:
    """Compute Levenshtein distance between two strings."""
    if len(a) < len(b):
        a, b = b, a
    if len(b) == 0:
        return len(a)
    prev_row = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr_row = [i]
        for j, cb in enumerate(b, 1):
            curr_row.append(min(
                prev_row[j] + 1,
                curr_row[j - 1] + 1,
                prev_row[j - 1] + (0 if ca == cb else 1),
            ))
        prev_row = curr_row
    return prev_row[-1]


def _are_phonetically_similar(word1: str, word2: str) -> bool:
    """Check if two words are phonetically/orthographically similar enough to be confusing.

    Words with edit distance <= 1 for short words (<= 5 chars) or edit distance <= 2
    for longer words that share the same first character are considered confusingly similar.
    """
    if not word1 or not word2:
        return False
    dist = _edit_distance(word1, word2)
    max_len = max(len(word1), len(word2))
    if max_len <= 5:
        return dist <= 1
    if dist <= 2 and word1[0] == word2[0]:
        return True
    return dist <= 1


def resequence_course_items(db: Session, user_id: UUID, course_id: UUID, *, commit: bool = True) -> None:
    """Auto-assign sort_order for all items in a course.

    Sequencing rules:
    1. Sort by difficulty_level ASC, then word length ASC, then alphabetical.
    2. Separate confusingly similar words (edit-distance check) by at least 2 positions.
    3. Assign sequential sort_order and unit_label based on difficulty tiers.

    commit=False is for callers mid-transaction (course-package import): an
    internal commit would persist a partially-imported package that the
    caller's error handling then claims was "not imported at all".
    """
    items = db.scalars(
        select(LearningItem).where(
            LearningItem.user_id == user_id,
            LearningItem.course_id == course_id,
        )
    ).all()

    if not items:
        return

    # Sort: difficulty_level ASC, then word length ASC, then alphabetical
    def sort_key(item: LearningItem) -> tuple[int, int, str]:
        clean_text = item.english_text.lower().strip()
        return (item.difficulty_level, len(clean_text), clean_text)

    sorted_items = sorted(items, key=sort_key)

    # Phonetic similarity grouping: avoid confusingly similar words back-to-back
    sequenced: list[LearningItem] = []
    remaining = list(sorted_items)
    while remaining:
        candidates = list(remaining)
        # Prefer the first remaining item that is NOT similar to the last sequenced item
        if sequenced:
            last_word = sequenced[-1].english_text.lower().strip()
            safe_candidates = [
                item for item in candidates
                if not _are_phonetically_similar(last_word, item.english_text.lower().strip())
            ]
            if safe_candidates:
                chosen = safe_candidates[0]
            else:
                chosen = candidates[0]
        else:
            chosen = candidates[0]
        remaining.remove(chosen)
        sequenced.append(chosen)

    # Assign sort_order and unit_label
    diff_unit_map: dict[int, str] = {
        1: "入门单元",
        2: "基础单元",
        3: "进阶单元",
        4: "提高单元",
        5: "挑战单元",
    }
    unit_counters: dict[int, int] = {1: 1, 2: 1, 3: 1, 4: 1, 5: 1}

    for idx, item in enumerate(sequenced):
        item.sort_order = idx
        level = item.difficulty_level
        if level not in unit_counters:
            level = 1
        unit_label = f"{diff_unit_map.get(level, '入门单元')} {unit_counters[level]}"
        item.unit_label = unit_label
        unit_counters[level] += 1

    if commit:
        db.commit()
    else:
        db.flush()
